"""
Analisador de Padrões - Sistema de Aprendizado de Regras de Expansão
Compara arquivo HB com arquivo de Referência para descobrir padrões automaticamente
"""

import openpyxl
from collections import defaultdict, Counter
import re

class AnalisadorPadroes:
    """Analisa padrões de expansão comparando HB com Referência"""
    
    def __init__(self, hb_path, ref_path):
        self.hb_path = hb_path
        self.ref_path = ref_path
        self.padroes_expansao = {}
        
    def carregar_hb(self):
        """Carrega dados do arquivo HB"""
        wb = openpyxl.load_workbook(self.hb_path)
        ws_acionamento = wb['Acionamento 1A']
        ws_status = wb['Status 1A']
        
        # Mapear acionamentos (cartao/anilha1 -> anilha2)
        acionamentos = []
        for idx, row in enumerate(ws_acionamento.iter_rows(min_row=2, values_only=True), start=2):
            if row[2] and row[3]:  # Tem cartão e anilha1
                acionamentos.append({
                    'linha_hb': idx,
                    'cartao': str(row[2]).strip(),
                    'anilha1': str(row[3]).strip(),
                    'anilha2': str(row[4]).strip() if row[4] else '',
                })
        
        # Mapear status (cartao/anilha1 -> descrição)
        status_map = {}
        for row in ws_status.iter_rows(min_row=2, values_only=True):
            if row[2] and row[3]:  # Tem cartão e anilha1
                key = f"{row[2]}|{row[3]}"
                status_map[key] = str(row[4]).strip() if row[4] else ''
        
        # Adicionar descrições aos acionamentos
        for acio in acionamentos:
            key = f"{acio['cartao']}|{acio['anilha1']}"
            acio['descricao'] = status_map.get(key, '')
        
        return acionamentos
    
    def carregar_referencia(self):
        """Carrega dados do arquivo de Referência"""
        wb = openpyxl.load_workbook(self.ref_path)
        ws = wb['Acionamento CCM-1A']
        
        referencias = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and str(row[0]).strip() != 'NOMENCLATURA':
                referencias.append({
                    'linha_ref': idx,
                    'nomenclatura': str(row[0]).strip(),
                    'descricao': str(row[1]).strip() if row[1] else '',
                    'cartao': str(row[2]).strip() if row[2] else '',
                    'anilha1': str(row[3]).strip() if row[3] else '',
                    'anilha2': str(row[4]).strip() if row[4] else '',
                    'tensao': str(row[5]).strip() if row[5] else '',
                })
        
        return referencias
    
    def correlacionar_dados(self, hb_list, ref_list):
        """Correlaciona dados HB com Referência usando cartão+anilha1 como chave"""
        
        # Agrupar referências por cartao+anilha1 (chave mais confiável)
        ref_por_chave = defaultdict(list)
        for ref in ref_list:
            if ref['anilha1']:
                # Normalizar cartão para comparação (remover espaços, maiúsculas, etc)
                cartao_norm = ref['cartao'].upper().replace(' ', '').replace('-', '')
                anilha1_norm = ref['anilha1'].upper().replace(' ', '')
                chave = f"{anilha1_norm}"  # Usar só anilha1 como chave principal
                ref_por_chave[chave].append(ref)
        
        # Também indexar por anilha2 para casos especiais
        ref_por_anilha2 = defaultdict(list)
        for ref in ref_list:
            if ref['anilha2']:
                anilha2_norm = ref['anilha2'].upper().replace(' ', '')
                ref_por_anilha2[anilha2_norm].append(ref)
        
        correlacoes = []
        for hb in hb_list:
            # Tentar correlacionar por anilha1 primeiro
            anilha1_norm = hb['anilha1'].upper().replace(' ', '')
            refs_correspondentes = ref_por_chave.get(anilha1_norm, [])
            
            # Se não encontrou, tentar por anilha2
            if not refs_correspondentes and hb['anilha2']:
                anilha2_norm = hb['anilha2'].upper().replace(' ', '')
                refs_correspondentes = ref_por_anilha2.get(anilha2_norm, [])
            
            if refs_correspondentes:
                correlacoes.append({
                    'hb': hb,
                    'refs': refs_correspondentes,
                    'fator_expansao': len(refs_correspondentes)
                })
        
        return correlacoes
    
    def extrair_padroes(self, correlacoes):
        """Extrai padrões de expansão das correlações"""
        
        padroes = defaultdict(lambda: {
            'count': 0,
            'fator_expansao': [],
            'exemplos': [],
            'nomenclaturas_geradas': set(),
            'estrutura_cartao': []
        })
        
        for corr in correlacoes:
            hb = corr['hb']
            refs = corr['refs']
            fator = corr['fator_expansao']
            
            # Identificar tipo baseado em descrição ou anilha
            tipo = self._identificar_tipo(hb)
            
            if tipo:
                padroes[tipo]['count'] += 1
                padroes[tipo]['fator_expansao'].append(fator)
                padroes[tipo]['exemplos'].append({
                    'hb_anilha2': hb['anilha2'],
                    'hb_descricao': hb['descricao'],
                    'ref_nomenclaturas': [r['nomenclatura'] for r in refs],
                    'ref_cartoes': [r['cartao'] for r in refs],
                })
                
                # Coletar nomenclaturas geradas
                for ref in refs:
                    padroes[tipo]['nomenclaturas_geradas'].add(ref['nomenclatura'])
                
                # Coletar estrutura de cartões
                for ref in refs:
                    padroes[tipo]['estrutura_cartao'].append(ref['cartao'])
        
        return dict(padroes)
    
    def _identificar_tipo(self, hb):
        """Identifica o tipo de equipamento baseado em descrição ou padrões"""
        desc = hb['descricao'].upper()
        anilha = hb['anilha2'].upper()
        
        # Padrões conhecidos
        if 'ATUADOR' in desc and 'RESERVA' not in desc:
            # Extrair número
            match = re.search(r'ATUADOR\s+(\d+)', desc)
            if match:
                return f"ATUADOR_{match.group(1)}"
            return "ATUADOR"
        
        if 'PISTÃO' in desc or 'PISTAO' in desc:
            match = re.search(r'PIST[ÃA]O\s+(\d+)', desc)
            if match:
                return f"PISTAO_{match.group(1)}"
            return "PISTAO"
        
        if 'MOTOR RESERVA' in desc or 'MOTOR DE RESERVA' in desc:
            match = re.search(r'RESERVA\s+(\d+)', desc)
            if match:
                return f"MOTOR_RESERVA_{match.group(1)}"
            return "MOTOR_RESERVA"
        
        if 'DESPELICULADORA' in desc:
            match = re.search(r'DESPELICULADORA\s+(\d+)', desc)
            if match:
                return f"DESPELICULADORA_{match.group(1)}"
            elif 'AUTORIZA' in desc:
                return "DESPELICULADORA_AUTORIZACAO"
            return "DESPELICULADORA"
        
        if 'SENSOR ELEVADOR' in desc or 'SENS' in desc and 'ELEVADOR' in desc:
            match = re.search(r'(\d+)', desc)
            if match:
                return f"SENSOR_ELEVADOR_{match.group(1)}"
            return "SENSOR_ELEVADOR"
        
        if 'INVERSOR' in desc:
            if 'RESERVA' in desc:
                return "INVERSOR_RESERVA"
            return "INVERSOR"
        
        if 'FOTOCELULA' in desc or 'FOTOC' in desc:
            return "FOTOCELULA"
        
        if 'VALVULA' in desc or 'VÁ LVULA' in desc:
            if 'GAS' in desc or 'GÁS' in desc:
                return "VALVULA_GAS"
            return "VALVULA"
        
        # Outros padrões por anilha
        if 'ACT-' in anilha:
            return "GENERICO_ACT"
        
        return None
    
    def analisar(self):
        """Executa análise completa e gera relatório"""
        print("=" * 100)
        print("ANALISADOR DE PADRÕES - SISTEMA DE APRENDIZADO")
        print("=" * 100)
        
        print("\n[1/4] Carregando arquivo HB...")
        hb_list = self.carregar_hb()
        print(f"      Total de acionamentos no HB: {len(hb_list)}")
        
        print("\n[2/4] Carregando arquivo de Referência...")
        ref_list = self.carregar_referencia()
        print(f"      Total de linhas na Referência: {len(ref_list)}")
        
        print("\n[3/4] Correlacionando dados HB <-> Referência...")
        correlacoes = self.correlacionar_dados(hb_list, ref_list)
        print(f"      Correlações encontradas: {len(correlacoes)}")
        
        print("\n[4/4] Extraindo padrões de expansão...")
        padroes = self.extrair_padroes(correlacoes)
        print(f"      Tipos de equipamentos identificados: {len(padroes)}")
        
        return padroes
    
    def gerar_relatorio(self, padroes):
        """Gera relatório detalhado dos padrões descobertos"""
        print("\n" + "=" * 100)
        print("PADRÕES DE EXPANSÃO DESCOBERTOS")
        print("=" * 100)
        
        for tipo, info in sorted(padroes.items()):
            print(f"\n{'─' * 100}")
            print(f"TIPO: {tipo}")
            print(f"{'─' * 100}")
            print(f"Ocorrências no HB: {info['count']}")
            
            # Fator de expansão
            fatores = info['fator_expansao']
            if fatores:
                fator_medio = sum(fatores) / len(fatores)
                print(f"Fator de Expansão: {fatores} (média: {fator_medio:.1f})")
            
            # Nomenclaturas geradas
            print(f"Nomenclaturas geradas: {sorted(info['nomenclaturas_geradas'])}")
            
            # Mostrar 3 primeiros exemplos
            print(f"\nExemplos ({min(3, len(info['exemplos']))}):")
            for i, ex in enumerate(info['exemplos'][:3], 1):
                print(f"  {i}. HB: {ex['hb_anilha2']} ({ex['hb_descricao'][:50]}...)")
                print(f"     -> REF: {ex['ref_nomenclaturas']}")
                if ex['ref_cartoes']:
                    cartoes_unicos = list(set([c for c in ex['ref_cartoes'] if c]))
                    if cartoes_unicos:
                        print(f"     -> Cartões REF: {cartoes_unicos}")
        
        print("\n" + "=" * 100)
        print("RESUMO DE REGRAS APRENDIDAS")
        print("=" * 100)
        
        for tipo, info in sorted(padroes.items()):
            fatores = info['fator_expansao']
            if fatores:
                fator_medio = sum(fatores) / len(fatores)
                print(f"{tipo:40s} -> {fator_medio:5.1f}x expansão ({info['count']} casos)")


if __name__ == "__main__":
    hb_path = r'HB - Zanchetta - Blancheamento 10.02.2025.xlsx'
    ref_path = r'Painel - referncia.xlsx'
    
    analisador = AnalisadorPadroes(hb_path, ref_path)
    padroes = analisador.analisar()
    analisador.gerar_relatorio(padroes)
    
    # Salvar padrões em arquivo para uso futuro
    import json
    with open('padroes_aprendidos.json', 'w', encoding='utf-8') as f:
        # Converter sets para listas para JSON
        padroes_json = {}
        for tipo, info in padroes.items():
            padroes_json[tipo] = {
                'count': info['count'],
                'fator_expansao': info['fator_expansao'],
                'nomenclaturas_geradas': list(info['nomenclaturas_geradas']),
                'exemplos': info['exemplos'][:5]  # Salvar só 5 primeiros exemplos
            }
        
        json.dump(padroes_json, f, indent=2, ensure_ascii=False)
    
    print("\n[OK] Padrões salvos em: padroes_aprendidos.json")
