"""
Gerador de Excel no formato Painel CCM
Responsável por criar o arquivo de saída formatado
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional
from dataclasses import dataclass
from datetime import datetime


@dataclass
class ColumnConfig:
    """Configuração de uma coluna"""
    name: str
    width: int = 15
    alignment: str = 'left'


class ExcelFormatter:
    """Formatador de estilos do Excel"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        
        # Estilos padrão
        self.header_style = NamedStyle(name='header_style')
        self.header_style.font = Font(bold=True, color='FFFFFF', size=11)
        self.header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.header_style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.data_style = NamedStyle(name='data_style')
        self.data_style.font = Font(size=10)
        self.data_style.alignment = Alignment(vertical='center')
        self.data_style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Larguras padrão de colunas
        self.column_widths = {
            'NOMENCLATURA': 18,
            'TIPO': 12,
            'DESCRICAO': 45,
            'CARTAO': 15,
            'ANILHA-CARTAO': 16,
            'ANILHA-RELE': 14,
            'RELE': 10,
            'CAVALO': 10,
            'BORNE': 10,
            'CABEAMENTO': 24,
            'FUSÍVEL': 12,
            'FUSIVEL': 12,
            'PAGINA REFERENCIA': 18,
            'INFORMACAO': 20,
            'VALOR': 50,
        }
    
    def apply_header_format(self, ws, row_num: int = 1):
        """Aplica formatação aos cabeçalhos"""
        for cell in ws[row_num]:
            cell.font = self.header_style.font
            cell.fill = self.header_style.fill
            cell.alignment = self.header_style.alignment
            cell.border = self.header_style.border
    
    def apply_data_format(self, ws, start_row: int = 2):
        """Aplica formatação às células de dados"""
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.font = self.data_style.font
                cell.alignment = self.data_style.alignment
                cell.border = self.data_style.border
    
    def adjust_column_widths(self, ws, columns: List[str] = None):
        """Ajusta largura das colunas"""
        for idx, col in enumerate(ws.columns, 1):
            col_letter = col[0].column_letter
            header_value = ws.cell(row=1, column=idx).value
            
            # Usa largura configurada ou calcula automaticamente
            if header_value and str(header_value).upper() in self.column_widths:
                width = self.column_widths[str(header_value).upper()]
            else:
                # Calcula baseado no conteúdo
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                width = min(max_length + 2, 50)
            
            ws.column_dimensions[col_letter].width = width
    
    def set_row_height(self, ws, height: int = 20):
        """Define altura das linhas"""
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = height


class PainelExcelGenerator:
    """Gerador principal do Excel no formato Painel CCM"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.formatter = ExcelFormatter(config.get('formatting', {}))
        self.wb: Optional[Workbook] = None
        self.painel_id: str = "1A"
    
    def generate(self, 
                 acionamentos: List[Dict],
                 status: List[Dict],
                 info_projeto: Dict,
                 output_path: str) -> bool:
        """
        Gera o arquivo Excel completo
        
        Args:
            acionamentos: Lista de dicionários com dados de acionamento
            status: Lista de dicionários com dados de status
            info_projeto: Dicionário com informações do projeto
            output_path: Caminho do arquivo de saída
        
        Returns:
            True se gerado com sucesso
        """
        try:
            self.wb = Workbook()
            self.painel_id = info_projeto.get('painel', '1A')
            
            # Remove aba padrão
            default_sheet = self.wb.active
            self.wb.remove(default_sheet)
            
            # Cria as abas na ordem correta
            self._create_descricao_sheet(acionamentos, status, info_projeto.get('paginas_referencia'))
            self._create_acionamento_sheet(acionamentos)
            self._create_reconhecimento_sheet(status)
            self._create_info_especiais_sheet(info_projeto)
            
            # Salva o arquivo
            self.wb.save(output_path)
            print(f"✅ Arquivo gerado com sucesso: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao gerar arquivo: {e}")
            return False
    
    def _create_descricao_sheet(self, acionamentos: List[Dict], status: List[Dict], 
                                   paginas_referencia: Dict[str, int] = None):
        """Cria aba de Descrição do Projeto"""
        ws = self.wb.create_sheet(f'Descrição de Projeto CCM-{self.painel_id}')
        
        # Cabeçalhos - C# espera exatamente estes nomes
        headers = ['NOMENCLATURA', 'DESCRICAO']
        ws.append(headers)
        
        # Itens especiais do projeto (páginas fixas)
        itens_especiais = [
            ('CAPA', 'Capa'),
            ('E-VISAO', 'E-VISAO'),
            ('I-VISAO', 'I-VISAO'),
            ('P-NOMENCLATURA', 'P-NOMENCLATURA'),
            ('COMANDO-1', 'COMANDO-1'),
            ('COMANDO-2', 'COMANDO-2'),
            ('COMANDO-3', 'COMANDO-3'),
            ('COMANDO-4', 'COMANDO-4'),
        ]
        
        # Adiciona itens especiais primeiro
        for nome, descricao in itens_especiais:
            ws.append([nome, descricao])
        
        # Coleta nomenclaturas únicas de acionamentos e status
        nomenclaturas = {}
        
        for item in acionamentos + status:
            nom = item.get('nomenclatura', '').strip()
            if nom and nom not in nomenclaturas:
                # Usa descrição se disponível
                descricao = item.get('descricao', nom)
                nomenclaturas[nom] = descricao
        
        # Adiciona nomenclaturas (ordenadas)
        for nom in sorted(nomenclaturas.keys()):
            descricao = nomenclaturas[nom]
            ws.append([nom, descricao])
        
        # Formatação
        self.formatter.apply_header_format(ws)
        self.formatter.apply_data_format(ws)
        self.formatter.adjust_column_widths(ws)
    
    def _create_acionamento_sheet(self, acionamentos: List[Dict]):
        """Cria aba de Acionamento"""
        ws = self.wb.create_sheet(f'Acionamento CCM-{self.painel_id}')
        
        # Cabeçalhos - EXATAMENTE como C# espera (sem acentos, ANILHA-RELE, FUSIVEL)
        headers = [
            'NOMENCLATURA', 'TIPO', 'DESCRICAO', 'CARTAO', 
            'ANILHA-CARTAO', 'ANILHA-RELE', 'RELE', 'CAVALO',
            'BORNE', 'CABEAMENTO', 'FUSIVEL'
        ]
        ws.append(headers)
        
        # Dados
        for item in acionamentos:
            # CV: converter para número ou deixar vazio (não "")
            cv_val = item.get('cv')
            if cv_val is None or cv_val == '' or (isinstance(cv_val, float) and pd.isna(cv_val)):
                cv_val = None
            
            row = [
                item.get('nomenclatura', ''),
                item.get('tipo') if item.get('tipo') else None,  # TIPO vazio se não definido
                item.get('descricao', ''),
                item.get('cartao', ''),
                item.get('anilha_cartao', ''),
                item.get('anilha_rele', ''),
                item.get('rele', ''),
                cv_val,
                item.get('borne', ''),
                item.get('cabeamento') if item.get('cabeamento') else None,
                item.get('fusivel', ''),
            ]
            ws.append(row)
        
        # Formatação
        self.formatter.apply_header_format(ws)
        self.formatter.apply_data_format(ws)
        self.formatter.adjust_column_widths(ws)
    
    def _create_reconhecimento_sheet(self, status: List[Dict]):
        """Cria aba de Reconhecimento (Status/DI)"""
        ws = self.wb.create_sheet(f'Reconhecimento CCM-{self.painel_id}')
        
        # Cabeçalhos - EXATAMENTE como C# espera (sem acentos)
        headers = [
            'NOMENCLATURA', 'TIPO', 'DESCRICAO', 'CARTAO',
            'ANILHA-CARTAO', 'BORNE', 'FUSIVEL'
        ]
        ws.append(headers)
        
        # Dados
        for item in status:
            row = [
                item.get('nomenclatura', ''),
                item.get('tipo') if item.get('tipo') and item.get('tipo') != 'STATUS' else None,
                item.get('descricao', ''),
                item.get('cartao', ''),
                item.get('anilha_cartao', ''),
                item.get('borne', ''),
                item.get('fusivel', ''),
            ]
            ws.append(row)
        
        # Formatação
        self.formatter.apply_header_format(ws)
        self.formatter.apply_data_format(ws)
        self.formatter.adjust_column_widths(ws)
    
    def _create_info_especiais_sheet(self, info_projeto: Dict):
        """Cria aba de Informações Especiais"""
        ws = self.wb.create_sheet(f'Informações Especiais CCM-{self.painel_id}')
        
        # Cabeçalhos
        headers = ['INFORMACAO', 'VALOR']
        ws.append(headers)
        
        # Informações padrão
        info_items = [
            ('local', info_projeto.get('local', '')),
            ('cliente', info_projeto.get('cliente', '')),
            ('nomeprojeto', info_projeto.get('nome_projeto', '')),
            ('desenhado', info_projeto.get('desenhado', 'Automático')),
            ('data', datetime.now().strftime('%d/%m/%Y')),
            ('conferido', info_projeto.get('conferido', '')),
            ('projeto_pagina', f"Sistema de Acionamento - Painel Centro de Controle de Motores - CCM-{self.painel_id}"),
        ]
        
        for key, value in info_items:
            ws.append([key, value])
        
        # Formatação
        self.formatter.apply_header_format(ws)
        self.formatter.apply_data_format(ws)
        self.formatter.adjust_column_widths(ws)


class ValidationReport:
    """Gerador de relatório de validação"""
    
    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.info: List[str] = []
    
    def add_error(self, message: str, row: int = None):
        """Adiciona um erro"""
        prefix = f"[Linha {row}] " if row else ""
        self.errors.append(f"{prefix}{message}")
    
    def add_warning(self, message: str, row: int = None):
        """Adiciona um aviso"""
        prefix = f"[Linha {row}] " if row else ""
        self.warnings.append(f"{prefix}{message}")
    
    def add_info(self, message: str):
        """Adiciona uma informação"""
        self.info.append(message)
    
    def has_errors(self) -> bool:
        """Verifica se há erros"""
        return len(self.errors) > 0
    
    def get_summary(self) -> str:
        """Retorna resumo do relatório"""
        lines = [
            "=" * 60,
            "RELATÓRIO DE VALIDAÇÃO",
            "=" * 60,
        ]
        
        if self.errors:
            lines.append(f"\n❌ ERROS ({len(self.errors)}):")
            for err in self.errors:
                lines.append(f"   • {err}")
        
        if self.warnings:
            lines.append(f"\n⚠️  AVISOS ({len(self.warnings)}):")
            for warn in self.warnings:
                lines.append(f"   • {warn}")
        
        if self.info:
            lines.append(f"\nℹ️  INFORMAÇÕES ({len(self.info)}):")
            for inf in self.info:
                lines.append(f"   • {inf}")
        
        if not self.errors and not self.warnings:
            lines.append("\n✅ Todos os dados validados com sucesso!")
        
        lines.append("\n" + "=" * 60)
        
        return '\n'.join(lines)
    
    def save_report(self, filepath: str):
        """Salva relatório em arquivo"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(self.get_summary())
